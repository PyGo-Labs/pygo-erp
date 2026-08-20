"""PyGo ERP V2.0 — Excel export."""
import os
import io
from datetime import datetime

base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
import sys
sys.path.insert(0, base)
sys.path.insert(0, os.path.join(base, "app"))

from core.registry import register


def get_db():
    """Use the request-scoped connection owned by core.main when available."""
    try:
        from core.main import get_db as _shared
        return _shared()
    except Exception:
        pass
    import sqlite3
    conn = sqlite3.connect(os.environ.get("PYGO_DB", "/tmp/pgerp.db"), timeout=15.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=15000")
    return conn


@register("core.export.excel")
def export_excel(table=None, **kwargs):
    """Export a table to Excel."""
    if not table:
        return {"error": "table required"}
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"error": "openpyxl not installed"}
    
    db = get_db()
    tables = [r["name"] for r in db.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE '_%' AND name NOT LIKE 'sqlite_%'"
    ).fetchall()]
    
    if table not in tables:
        return {"error": f"table {table} not found", "available": tables}
    
    rows = db.execute(f"SELECT * FROM {table}").fetchall()
    if not rows:
        return {"error": "no data"}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table
    
    headers = list(rows[0].keys())
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[header])
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return {
        "table": table,
        "format": "xlsx",
        "rows": len(rows),
        "content": output.getvalue(),
        "filename": f"{table}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
    }
